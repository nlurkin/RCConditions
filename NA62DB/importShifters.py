#!/bin/env python

'''
Created on Apr 27, 2016

@author: nlurkin
'''

import re
import sys

from openpyxl import load_workbook

from DBConfig import DBConfig as DB
from database import DBConnector


if __name__ == '__main__':
    wb = load_workbook("participants-list.xlsx")

    myconn = DBConnector(True)
    myconn.connectDB(passwd=sys.argv[1], host=DB.host, db="shifters", user=DB.userName, port=DB.port)
    
    # search name
    for row in range(8,1000):
        cellVal = wb["Sheet0"].cell(row=row, column=2).value
        if cellVal==None:
            break
        m = re.search("([A-Z][a-z\s]*) ([A-Z\s]*)", cellVal)
        if m:
            (name, surname) = m.groups()
        
        r = myconn.executeGet("SELECT COUNT(*) FROM shifter WHERE name LIKE %s AND surname LIKE %s", [name, surname])
        if r[0]["COUNT(*)"]==0:
            myconn.executeInsert("INSERT INTO shifter (name, surname) VALUES (%s,%s)", [name, surname])